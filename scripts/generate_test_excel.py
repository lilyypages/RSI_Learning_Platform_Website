#!/usr/bin/env python3
"""Generate test documentation Excel file for RSI Learning Platform"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style definitions ──
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL_AUTH = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FILL_PRINCIPAL = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
HEADER_FILL_TEACHER = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
HEADER_FILL_STUDENT = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
HEADER_FILL_PARENT = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
HEADER_FILL_CROSS = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
HEADER_FILL_MSG = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
HEADER_FILL_FE = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

SECTION_FONT = Font(name="Arial", bold=True, size=12, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

PASS_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")


def add_section_header(ws, row, text, fill, col_count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    return row + 1


def add_headers(ws, row, headers, fill):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24
    return row + 1


def add_row(ws, row, values):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = DATA_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    return row + 1


# ═══════════════════════════════════════════════════════════════
# SHEET 1: BACKEND TEST
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Backend Test"
ws1.sheet_properties.tabColor = "27AE60"

BE_HEADERS = ["No", "Test ID", "Endpoint", "Method", "Role", "Request Body", "Expected Status", "Expected Result", "Actual Result", "Status"]
col_widths = [5, 8, 35, 8, 12, 45, 10, 40, 30, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

row = 1
# Title
ws1.merge_cells("A1:J1")
title = ws1.cell(row=1, column=1, value="RSI Learning Platform — Backend API Test (4 Roles)")
title.font = Font(name="Arial", bold=True, size=14, color="2E7D32")
title.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32
row = 3

# A. AUTH
row = add_section_header(ws1, row, "A. AUTHENTICATION (Login & Logout)", HEADER_FILL_AUTH, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_AUTH)
auth_tests = [
    [1, "A01", "/api/auth/login", "POST", "PRINCIPAL", '{"email":"kepsek@test.com","password":"admin123"}', 200, '{success:true, role:"PRINCIPAL"}', "", "⬜"],
    [2, "A02", "/api/auth/login", "POST", "TEACHER", '{"email":"guru@test.com","password":"admin123"}', 200, '{success:true, role:"TEACHER"}', "", "⬜"],
    [3, "A03", "/api/auth/login", "POST", "STUDENT", '{"email":"siswa@test.com","password":"admin123"}', 200, '{success:true, role:"STUDENT"}', "", "⬜"],
    [4, "A04", "/api/auth/login", "POST", "PARENT", '{"email":"ortu@test.com","password":"admin123"}', 200, '{success:true, role:"PARENT"}', "", "⬜"],
    [5, "A05", "/api/auth/login", "POST", "—", '{"email":"kepsek@test.com","password":"wrongpass"}', 401, '{success:false} password salah', "", "⬜"],
    [6, "A06", "/api/auth/login", "POST", "—", '{"email":"notanemail"}', 400, '{success:false} validation error', "", "⬜"],
    [7, "H01", "/api/auth/logout", "POST", "PRINCIPAL", "—", 200, "Session dihapus", "", "⬜"],
    [8, "H02", "/api/auth/logout", "POST", "TEACHER", "—", 200, "Session dihapus", "", "⬜"],
    [9, "H03", "/api/auth/logout", "POST", "STUDENT", "—", 200, "Session dihapus", "", "⬜"],
    [10, "H04", "/api/auth/logout", "POST", "PARENT", "—", 200, "Session dihapus", "", "⬜"],
]
for t in auth_tests:
    row = add_row(ws1, row, t)
row += 1

# B. PRINCIPAL
row = add_section_header(ws1, row, "B. PRINCIPAL (Kepala Sekolah) Endpoints", HEADER_FILL_PRINCIPAL, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_PRINCIPAL)
principal_tests = [
    [11, "B01", "/api/users", "GET", "PRINCIPAL", "—", 200, "Array of users [{id,name,email,role}]", "", "⬜"],
    [12, "B02", "/api/teachers", "GET", "PRINCIPAL", "—", 200, "{success:true, teachers:[...]}', ", "", "⬜"],
    [13, "B03", "/api/students", "GET", "PRINCIPAL", "—", 200, "{success:true, students:[...]}", "", "⬜"],
    [14, "B04", "/api/classes", "GET", "PRINCIPAL", "—", 200, "Array of classes", "", "⬜"],
    [15, "B05", "/api/subjects", "GET", "PRINCIPAL", "—", 200, "Array of subjects", "", "⬜"],
    [16, "B06", "/api/audit-log", "GET", "PRINCIPAL", "—", 200, "Array of audit entries", "", "⬜"],
    [17, "B07", "/api/progress", "GET", "PRINCIPAL", "—", 200, "Progress overview semua siswa", "", "⬜"],
    [18, "B08", "/api/profile", "GET", "PRINCIPAL", "—", 200, '{userId, name, role:"PRINCIPAL"}', "", "⬜"],
    [19, "B09", "/api/notifications", "GET", "PRINCIPAL", "—", 200, "Array of notifications", "", "⬜"],
    [20, "B10", "/api/auth/reset-password", "POST", "PRINCIPAL", '{"userId":"<target>"}', "200/404", "Success or user not found", "", "⬜"],
]
for t in principal_tests:
    row = add_row(ws1, row, t)
row += 1

# C. TEACHER
row = add_section_header(ws1, row, "C. TEACHER (Guru) Endpoints", HEADER_FILL_TEACHER, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_TEACHER)
teacher_tests = [
    [21, "C01", "/api/guru/dashboard", "GET", "TEACHER", "—", 200, "{success:true, stats:{...}}", "", "⬜"],
    [22, "C02", "/api/guru/subjects", "GET", "TEACHER", "—", 200, "Array of classSubjects", "", "⬜"],
    [23, "C03", "/api/materials", "GET", "TEACHER", "—", 200, "Array of materials", "", "⬜"],
    [24, "C04", "/api/students?includeProgress=true", "GET", "TEACHER", "—", 200, "{success:true, students:[...]}", "", "⬜"],
    [25, "C05", "/api/reports", "GET", "TEACHER", "—", 200, "Array of weekly reports", "", "⬜"],
    [26, "C06", "/api/messages", "GET", "TEACHER", "—", 200, "Array of messages", "", "⬜"],
    [27, "C07", "/api/questions", "GET", "TEACHER", "—", 200, "Array of questions", "", "⬜"],
    [28, "C08", "/api/profile", "GET", "TEACHER", "—", 200, '{userId, name, role:"TEACHER"}', "", "⬜"],
    [29, "C09", "/api/notifications", "GET", "TEACHER", "—", 200, "Array of notifications", "", "⬜"],
    [30, "C10", "/api/progress", "GET", "TEACHER", "—", 200, "Progress data siswa", "", "⬜"],
]
for t in teacher_tests:
    row = add_row(ws1, row, t)
row += 1

# D. STUDENT
row = add_section_header(ws1, row, "D. STUDENT (Siswa) Endpoints", HEADER_FILL_STUDENT, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_STUDENT)
student_tests = [
    [31, "D01", "/api/profile", "GET", "STUDENT", "—", 200, '{userId, name, role:"STUDENT", teachers:[...]}', "", "⬜"],
    [32, "D02", "/api/progress", "GET", "STUDENT", "—", 200, "Progress belajar per mapel", "", "⬜"],
    [33, "D03", "/api/materials", "GET", "STUDENT", "—", 200, "Array of published materials", "", "⬜"],
    [34, "D04", "/api/messages", "GET", "STUDENT", "—", 200, "Array of messages dari guru", "", "⬜"],
    [35, "D05", "/api/quiz/sessions", "GET", "STUDENT", "—", 200, "Array of quiz sessions", "", "⬜"],
    [36, "D06", "/api/notifications", "GET", "STUDENT", "—", 200, "Array of notifications", "", "⬜"],
    [37, "D07", "/api/subjects", "GET", "STUDENT", "—", 200, "Array of subjects", "", "⬜"],
]
for t in student_tests:
    row = add_row(ws1, row, t)
row += 1

# E. PARENT
row = add_section_header(ws1, row, "E. PARENT (Orang Tua) Endpoints", HEADER_FILL_PARENT, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_PARENT)
parent_tests = [
    [38, "E01", "/api/profile", "GET", "PARENT", "—", 200, '{userId, name, role:"PARENT", students:[...]}', "", "⬜"],
    [39, "E02", "/api/reports", "GET", "PARENT", "—", 200, "Array of weekly reports (anak)", "", "⬜"],
    [40, "E03", "/api/messages", "GET", "PARENT", "—", 200, "Messages termasuk forwarded 💬", "", "⬜"],
    [41, "E04", "/api/notifications", "GET", "PARENT", "—", 200, "Array of notifications", "", "⬜"],
    [42, "E05", "/api/progress", "GET", "PARENT", "—", 200, "Progress belajar anak", "", "⬜"],
]
for t in parent_tests:
    row = add_row(ws1, row, t)
row += 1

# F. CROSS-ROLE
row = add_section_header(ws1, row, "F. CROSS-ROLE AUTHORIZATION (Negative Tests)", HEADER_FILL_CROSS, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_CROSS)
cross_tests = [
    [43, "F01", "/api/guru/dashboard", "GET", "STUDENT→TEACHER", "—", 403, '{success:false, message:"Forbidden"}', "", "⬜"],
    [44, "F02", "/api/guru/dashboard", "GET", "PARENT→TEACHER", "—", 403, '{success:false, message:"Forbidden"}', "", "⬜"],
    [45, "F03", "/api/audit-log", "GET", "TEACHER→PRINCIPAL", "—", 403, "{success:false} hanya kepsek", "", "⬜"],
    [46, "F04", "/api/audit-log", "GET", "STUDENT→PRINCIPAL", "—", 403, "{success:false}", "", "⬜"],
    [47, "F05", "/api/profile", "GET", "No Auth", "—", 401, '{success:false, message:"Unauthorized"}', "", "⬜"],
]
for t in cross_tests:
    row = add_row(ws1, row, t)
row += 1

# G. MESSAGING
row = add_section_header(ws1, row, "G. MESSAGING & SYNC (Write Operations)", HEADER_FILL_MSG, 10)
row = add_headers(ws1, row, BE_HEADERS, HEADER_FILL_MSG)
msg_tests = [
    [48, "G01", "/api/messages", "POST", "TEACHER", '{"receiverId":"<siswaId>","content":"Halo..."}', 201, "Pesan tersimpan + forward ke parent", "", "⬜"],
    [49, "G02", "/api/messages", "POST", "STUDENT", '{"receiverId":"<guruId>","content":"Terima kasih!"}', 201, "Pesan tersimpan + notifikasi", "", "⬜"],
    [50, "G03", "/api/messages", "GET", "PARENT", "—", 200, "Pesan forwarded 💬 muncul di inbox", "", "⬜"],
    [51, "G04", "/api/reports", "POST", "TEACHER", '{"classSubjectId","catatanKelas","laporanSiswa"}', 201, "Laporan + pesan 📋 ke parent", "", "⬜"],
    [52, "G05", "/api/chat", "POST", "TEACHER", '{"receiverId":"<siswaId>","content":"..."}', 201, "Pesan + forward ke parent", "", "⬜"],
]
for t in msg_tests:
    row = add_row(ws1, row, t)


# ═══════════════════════════════════════════════════════════════
# SHEET 2: FRONTEND TEST
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Frontend Test")
ws2.sheet_properties.tabColor = "2980B9"

FE_HEADERS = ["No", "Test ID", "Halaman / URL", "Role", "Deskripsi", "Expected Behavior", "Actual Result", "Status"]
fe_widths = [5, 8, 40, 12, 35, 45, 30, 10]
for i, w in enumerate(fe_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

row = 1
ws2.merge_cells("A1:H1")
title2 = ws2.cell(row=1, column=1, value="RSI Learning Platform — Frontend UI Test (4 Roles)")
title2.font = Font(name="Arial", bold=True, size=14, color="2E7D32")
title2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32
row = 3

# Login & Routing
row = add_section_header(ws2, row, "H. LOGIN & ROUTING", HEADER_FILL_AUTH, 8)
row = add_headers(ws2, row, FE_HEADERS, HEADER_FILL_AUTH)
login_fe = [
    [53, "FE01", "/auth/login", "—", "Halaman login tampil", "Form email + password, tombol login", "", "⬜"],
    [54, "FE02", "/auth/login → /dashboard/kepsek", "PRINCIPAL", "Login kepsek redirect", "Redirect ke /dashboard/kepsek", "", "⬜"],
    [55, "FE03", "/auth/login → /dashboard/guru", "TEACHER", "Login guru redirect", "Redirect ke /dashboard/guru", "", "⬜"],
    [56, "FE04", "/auth/login → /dashboard/siswa", "STUDENT", "Login siswa redirect", "Redirect ke /dashboard/siswa", "", "⬜"],
    [57, "FE05", "/auth/login → /dashboard/ortu", "PARENT", "Login ortu redirect", "Redirect ke /dashboard/ortu", "", "⬜"],
]
for t in login_fe:
    row = add_row(ws2, row, t)
row += 1

# PRINCIPAL FE
row = add_section_header(ws2, row, "I. PRINCIPAL (Kepsek) Dashboard", HEADER_FILL_PRINCIPAL, 8)
row = add_headers(ws2, row, FE_HEADERS, HEADER_FILL_PRINCIPAL)
kepsek_fe = [
    [58, "FE06", "/dashboard/kepsek", "PRINCIPAL", "Dashboard utama", "Statistik, grafik, jumlah guru/siswa", "", "⬜"],
    [59, "FE07", "/dashboard/kepsek/guru", "PRINCIPAL", "Manajemen guru", "Tabel daftar guru, tombol tambah", "", "⬜"],
    [60, "FE08", "/dashboard/kepsek/guru/tambah", "PRINCIPAL", "Form tambah guru", "Form input + validasi", "", "⬜"],
    [61, "FE09", "/dashboard/kepsek/guru/[id]", "PRINCIPAL", "Detail guru", "Info guru, mapel yang diajar", "", "⬜"],
    [62, "FE10", "/dashboard/kepsek/guru/[id]/edit", "PRINCIPAL", "Edit data guru", "Form edit + simpan", "", "⬜"],
    [63, "FE11", "/dashboard/kepsek/guru/[id]/mapel", "PRINCIPAL", "Kelola mapel guru", "Assign/unassign mapel", "", "⬜"],
    [64, "FE12", "/dashboard/kepsek/siswa", "PRINCIPAL", "Manajemen siswa", "Tabel daftar siswa", "", "⬜"],
    [65, "FE13", "/dashboard/kepsek/siswa/tambah", "PRINCIPAL", "Form tambah siswa", "Form input + validasi", "", "⬜"],
    [66, "FE14", "/dashboard/kepsek/siswa/[id]", "PRINCIPAL", "Detail siswa", "Info siswa, kelas, progress", "", "⬜"],
    [67, "FE15", "/dashboard/kepsek/siswa/[id]/edit", "PRINCIPAL", "Edit data siswa", "Form edit + simpan", "", "⬜"],
    [68, "FE16", "/dashboard/kepsek/siswa/[id]/progress", "PRINCIPAL", "Progress siswa", "Grafik progress per mapel", "", "⬜"],
    [69, "FE17", "/dashboard/kepsek/audit", "PRINCIPAL", "Log audit", "Tabel log aktivitas sistem", "", "⬜"],
]
for t in kepsek_fe:
    row = add_row(ws2, row, t)
row += 1

# TEACHER FE
row = add_section_header(ws2, row, "J. TEACHER (Guru) Dashboard", HEADER_FILL_TEACHER, 8)
row = add_headers(ws2, row, FE_HEADERS, HEADER_FILL_TEACHER)
guru_fe = [
    [70, "FE18", "/dashboard/guru", "TEACHER", "Dashboard utama guru", "Statistik mengajar, kelas", "", "⬜"],
    [71, "FE19", "/dashboard/guru/mapel", "TEACHER", "Daftar mapel diajar", "Card mapel yang di-assign", "", "⬜"],
    [72, "FE20", "/dashboard/guru/mapel/input-materi", "TEACHER", "Input materi baru", "Form judul, konten, video", "", "⬜"],
    [73, "FE21", "/dashboard/guru/mapel/kelola-materi", "TEACHER", "Kelola materi", "List, edit, delete, publish", "", "⬜"],
    [74, "FE22", "/dashboard/guru/mapel/kelola-soal", "TEACHER", "Kelola soal quiz", "Bank soal, CRUD", "", "⬜"],
    [75, "FE23", "/dashboard/guru/monitoring", "TEACHER", "Monitoring siswa", "Daftar siswa, progress", "", "⬜"],
    [76, "FE24", "/dashboard/guru/monitoring/[id]", "TEACHER", "Detail monitoring", "Progress individual", "", "⬜"],
    [77, "FE25", "/dashboard/guru/chat", "TEACHER", "Chat guru ↔ siswa", "Pilih siswa, kirim/terima pesan", "", "⬜"],
    [78, "FE26", "/dashboard/guru/laporan", "TEACHER", "Kirim laporan mingguan", "Form catatan kelas + individu", "", "⬜"],
    [79, "FE27", "/dashboard/guru/profile", "TEACHER", "Profil guru", "Info guru, tombol edit", "", "⬜"],
]
for t in guru_fe:
    row = add_row(ws2, row, t)
row += 1

# STUDENT FE
row = add_section_header(ws2, row, "K. STUDENT (Siswa) Dashboard", HEADER_FILL_STUDENT, 8)
row = add_headers(ws2, row, FE_HEADERS, HEADER_FILL_STUDENT)
siswa_fe = [
    [80, "FE28", "/dashboard/siswa", "STUDENT", "Dashboard utama siswa", "Hero, stats, progress mapel, quiz", "", "⬜"],
    [81, "FE29", "/dashboard/siswa/mapel", "STUDENT", "Daftar materi/mapel", "Card mapel yang diikuti", "", "⬜"],
    [82, "FE30", "/dashboard/siswa/mapel/[mapelId]", "STUDENT", "Detail mapel", "Daftar materi, progress bar", "", "⬜"],
    [83, "FE31", "/dashboard/siswa/materials/[id]", "STUDENT", "Detail materi", "Konten, video, tombol quiz", "", "⬜"],
    [84, "FE32", "/dashboard/siswa/belajar/[materiId]", "STUDENT", "Halaman belajar", "Konten materi interaktif", "", "⬜"],
    [85, "FE33", "/dashboard/siswa/quiz", "STUDENT", "Riwayat quiz", "Daftar quiz selesai", "", "⬜"],
    [86, "FE34", "/dashboard/siswa/quiz/[materialId]", "STUDENT", "Kerjakan quiz", "Soal quiz adaptif, submit", "", "⬜"],
    [87, "FE35", "/dashboard/siswa/pesan", "STUDENT", "Pesan siswa ↔ guru", "Inbox, compose, reply", "", "⬜"],
]
for t in siswa_fe:
    row = add_row(ws2, row, t)
row += 1

# PARENT FE
row = add_section_header(ws2, row, "L. PARENT (Orang Tua) Dashboard", HEADER_FILL_PARENT, 8)
row = add_headers(ws2, row, FE_HEADERS, HEADER_FILL_PARENT)
ortu_fe = [
    [88, "FE36", "/dashboard/ortu", "PARENT", "Ringkasan anak", "Info anak, stats, progress", "", "⬜"],
    [89, "FE37", "/dashboard/ortu/grafik", "PARENT", "Grafik kemajuan", "Chart progress per mapel/waktu", "", "⬜"],
    [90, "FE38", "/dashboard/ortu/pesan", "PARENT", "Kotak pesan guru", "Inbox + forwarded messages 💬", "", "⬜"],
    [91, "FE39", "/dashboard/ortu/profile", "PARENT", "Profil & akun", "Info profil, tombol edit", "", "⬜"],
]
for t in ortu_fe:
    row = add_row(ws2, row, t)
row += 1

# SYNC FE
row = add_section_header(ws2, row, "M. SINKRONISASI GURU → PARENT", HEADER_FILL_MSG, 8)
row = add_headers(ws2, row, FE_HEADERS, HEADER_FILL_MSG)
sync_fe = [
    [92, "FE40", "Guru chat → siswa → parent", "TEACHER→PARENT", "Pesan guru forward ke ortu", "Pesan 💬 muncul di inbox siswa & parent", "", "⬜"],
    [93, "FE41", "Guru laporan → parent", "TEACHER→PARENT", "Laporan forward ke ortu", "📋 Laporan Mingguan masuk inbox parent", "", "⬜"],
    [94, "FE42", "Parent balas pesan", "PARENT→TEACHER", "Ortu reply ke guru", "Pesan terkirim + notifikasi guru", "", "⬜"],
    [95, "FE43", "Sidebar siswa", "STUDENT", "Navigasi sidebar", "Beranda, Materi, Quiz, Pesan Guru", "", "⬜"],
]
for t in sync_fe:
    row = add_row(ws2, row, t)


# ═══════════════════════════════════════════════════════════════
# SHEET 3: SUMMARY
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Summary")
ws3.sheet_properties.tabColor = "2E7D32"

ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 15
ws3.column_dimensions["C"].width = 25

ws3.merge_cells("A1:C1")
t3 = ws3.cell(row=1, column=1, value="Test Coverage Summary")
t3.font = Font(name="Arial", bold=True, size=14, color="2E7D32")
t3.alignment = Alignment(horizontal="center")

row = 3
sum_headers = ["Kategori", "Jumlah Test", "Role Coverage"]
row = add_headers(ws3, row, sum_headers, PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"))

summary = [
    ["Auth (Login/Logout)", 10, "Semua 4 role + negative"],
    ["Principal Backend", 10, "PRINCIPAL"],
    ["Teacher Backend", 10, "TEACHER"],
    ["Student Backend", 7, "STUDENT"],
    ["Parent Backend", 5, "PARENT"],
    ["Cross-Role Authorization", 5, "Mixed (negative tests)"],
    ["Messaging & Sync", 5, "TEACHER → STUDENT → PARENT"],
    ["Frontend Login/Routing", 5, "Semua 4 role"],
    ["Frontend Principal", 12, "PRINCIPAL"],
    ["Frontend Teacher", 10, "TEACHER"],
    ["Frontend Student", 8, "STUDENT"],
    ["Frontend Parent", 4, "PARENT"],
    ["Frontend Sync", 4, "TEACHER ↔ PARENT"],
]
total = 0
for s in summary:
    row = add_row(ws3, row, s)
    total += s[1]

row += 1
total_cell = ws3.cell(row=row, column=1, value="TOTAL")
total_cell.font = Font(name="Arial", bold=True, size=12)
total_val = ws3.cell(row=row, column=2, value=total)
total_val.font = Font(name="Arial", bold=True, size=12, color="2E7D32")
ws3.cell(row=row, column=3, value="4 roles fully covered").font = Font(name="Arial", bold=True, size=10, color="2E7D32")

# Save
OUTPUT = "/home/archian/RSI_Learning_Platform_Website/docs/Test_Script_RSI_Learning_Platform.xlsx"
import os
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)
print(f"✅ Excel file saved: {OUTPUT}")
